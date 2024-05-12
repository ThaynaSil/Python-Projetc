from tkinter import * #Importando biblioteca para a criação do formulario.
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import threading

dados_cache = []
ultimo_salvamento = datetime.now()
thread_running = True


def limpar_campos():
    entry_Nome.delete(0, 'end')
    entry_Idade.delete(0, 'end')
    entry_Mat.delete(0, 'end')
    combo_Status.set("")

def submit_form():
    # Função para coletar os dados dos campos e fazer algo com eles
    nome = entry_Nome.get()
    idade = entry_Idade.get()
    matricula = entry_Mat.get()
    status = combo_Status.get()

    dados_cache.append((nome, idade, matricula, status))

    limpar_campos()


def inserir_dados_no_excel(lista):
    # Verifique se o arquivo Excel já existe
    if os.path.exists("dados.xlsx"):
        # Se o arquivo existir, carregue-o
        workbook = load_workbook("dados.xlsx")
        sheet = workbook.active
    else:
        # Se o arquivo não existir, crie um novo
        workbook = Workbook()
        sheet = workbook.active
        # Adicione cabeçalhos
        sheet["A1"] = "Nome"
        sheet["B1"] = "Idade"
        sheet["C1"] = "Matricula"
        sheet["D1"] = "Status"


    # Insira os dados na próxima linha
    for dado in lista:
        nome, idade, matricula, status = dado
        proxima_linha = sheet.max_row + 1
        sheet[f"A{proxima_linha}"] = nome
        sheet[f"B{proxima_linha}"] = idade
        sheet[f"C{proxima_linha}"] = matricula
        sheet[f"D{proxima_linha}"] = status

    workbook.save("dados.xlsx")


janela = Tk()

janela.title("Formulário")

# Ajustar o tamanho da janela
janela.geometry("300x250")  # Largura x Altura

janela.configure(background='#9BA4B8')

# Criação dos Textos e campos de entrada de dados
Texto_Nome = Label(janela, text="Nome:", font=("Arial", 10))
Texto_Nome.grid(column=0, row=0, sticky="w", pady= 10, padx=5) #Posição Do Texto
entry_Nome = Entry(janela) 
entry_Nome.grid(column=1, row=0, sticky="w", pady= 10) #Posicao Da Entrada

Texto_Idade = Label(janela, text="Idade:", font=("Arial", 10))
Texto_Idade.grid(column=0, row=1, sticky="w", pady= 10, padx=5)
entry_Idade = Entry(janela)
entry_Idade.grid(column=1, row=1, sticky="w", pady= 10)

Texto_Mat = Label(janela, text="Matrícula:", font=("Arial", 10))
Texto_Mat.grid(column=0, row=2, sticky="w", pady= 10, padx=5)
entry_Mat = Entry(janela)
entry_Mat.grid(column=1, row=2, sticky="w", pady= 10)

Texto_Status = Label(janela, text="Status:", font=("Arial", 10))
Texto_Status.grid(column=0, row=3, sticky="w", pady=10, padx=5)
# Usando Combobox para as opções de status
status_options = ["Ativo", "Inativo"]
combo_Status = ttk.Combobox(janela, values=status_options, state="readonly")
combo_Status.grid(column=1, row=3, sticky="w", pady=10)


# Botão para enviar o formulário
submit_button = Button(janela, text=" Enviar ", font=("Arial", 11), command=submit_form) #Criação do botao - Quando clicado chama a função
submit_button.grid(columnspan=5, row=7, pady=20) # Posição na janela

def salvar_periodicamente():
    global ultimo_salvamento, thread_running

    while thread_running:
        tempo_passado = (datetime.now() - ultimo_salvamento).seconds
        if tempo_passado >= 60:
            # Salve os dados em uma planilha do Excel
            inserir_dados_no_excel(dados_cache)
            # Limpe a lista de cache
            dados_cache.clear()
            # Atualize o tempo do último salvamento
            ultimo_salvamento = datetime.now()

threading.Thread(target=salvar_periodicamente).start()

def fechar_janela():
    global thread_running
    thread_running = False
    janela.destroy()

# Associe a função fechar_janela ao evento de fechar a janela
janela.protocol("WM_DELETE_WINDOW", fechar_janela)

janela.mainloop()